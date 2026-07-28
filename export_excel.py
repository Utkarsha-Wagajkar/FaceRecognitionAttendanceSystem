import sqlite3
import os
from openpyxl import Workbook
from openpyxl.styles import Font

# -----------------------------
# Database Path
# -----------------------------
DB_PATH = os.path.join("database", "attendance.db")

# -----------------------------
# Export Folder
# -----------------------------
EXPORT_FOLDER = "exports"

# Create folder if not exists
os.makedirs(EXPORT_FOLDER, exist_ok=True)

# -----------------------------
# Connect Database
# -----------------------------
conn = sqlite3.connect(DB_PATH)
cursor = conn.cursor()

# -----------------------------
# Fetch Attendance Records
# -----------------------------
cursor.execute("""
SELECT
    attendance.student_id,
    students.name,
    students.department,
    attendance.date,
    attendance.time,
    attendance.status
FROM attendance
INNER JOIN students
ON attendance.student_id = students.student_id
ORDER BY attendance.date DESC, attendance.time DESC
""")

records = cursor.fetchall()

conn.close()

# -----------------------------
# Create Workbook
# -----------------------------
workbook = Workbook()

sheet = workbook.active

sheet.title = "Attendance"

# -----------------------------
# Header
# -----------------------------
headers = [
    "Student ID",
    "Name",
    "Department",
    "Date",
    "Time",
    "Status"
]

for column, header in enumerate(headers, start=1):

    cell = sheet.cell(row=1, column=column)

    cell.value = header

    cell.font = Font(bold=True)

# -----------------------------
# Write Data
# -----------------------------
for row_index, row_data in enumerate(records, start=2):

    for column_index, value in enumerate(row_data, start=1):

        sheet.cell(
            row=row_index,
            column=column_index
        ).value = value

# -----------------------------
# Adjust Column Width
# -----------------------------
for column_cells in sheet.columns:

    length = max(len(str(cell.value)) if cell.value else 0 for cell in column_cells)

    sheet.column_dimensions[column_cells[0].column_letter].width = length + 5

# -----------------------------
# Save File
# -----------------------------
file_path = os.path.join(
    EXPORT_FOLDER,
    "attendance.xlsx"
)

workbook.save(file_path)

print("✅ Attendance exported successfully!")
print("Saved at:", file_path)